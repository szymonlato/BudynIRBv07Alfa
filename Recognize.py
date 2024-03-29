import threading
from datetime import datetime, date, timedelta
from tkinter import *
import numpy as np
import openpyxl
import pandas as pd
import cv2 as cv
import time
from PyQt5 import uic
from PyQt5.QtCore import QTimer, pyqtSignal
from PyQt5.QtGui import QFont
from PyQt5.QtWidgets import *
from PlikiUi.OknoLogowania_ui import *
from PlikiUi.OknoGlowne_ui import *
from PlikiUi.OknoAdministratora_ui import *
from PlikiUi.OknoGeneratoraHasel_ui import *
import sys
import pymssql
import hashlib
import os
from pathlib import Path
import shutil
from openpyxl import Workbook
from openpyxl.styles import Alignment
import string
import random
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Spacer
from reportlab.lib import colors

# python -m PyQt5.uic.pyuic -x PlikiUi/OknoGlowne.ui -o PlikiUi/OknoGlowne_ui.py
# C:\Users\kamil\AppData\Local\Programs\Python\Python39\Lib\site-packages\QtDesigner

server = 'dbserver-projekt-zespolowy-dt.database.windows.net'
database = 'db'
username = 'Ladybug'
password = 'FfqGV3PY'

nazwa_usera = ""
id_names = pd.read_csv('id_names.csv')
id_names = id_names[['id', 'name']]
faceClassifier = cv.CascadeClassifier('Classifiers/haarface.xml')
lbph = cv.face.LBPHFaceRecognizer_create(threshold=500)
lbph.read('Classifiers/TrainedLBPH.yml')

powiadomienieOKoncuPrzepustki = ["05.00.00", "05.15.00"]
powiadomienieOStarejSluzbie = ["09.00.00"]
archiwizacjaDokumentu = ["01.00.00"]

pathToIRB = "Dokumenty/Imienne_Rozliczenie_Bojowe.xlsx"
pathToPrzepustki = "Dokumenty/Przepustki.xlsx"
pathToNumer = "Dokumenty/Spis_Numerow.txt"
pathToSluzba = "Dokumenty/Lista_Sluzb.xlsx"
pathToArchive = "ZarchiwizowaneDokumenty"

def EdycjaBazyDanych(osobaDoZmiany, nowyPowod):
    server = 'dbserver-projekt-zespolowy-dt.database.windows.net'
    database = 'db'
    username = 'Ladybug'
    password = 'FfqGV3PY'
    conn = pymssql.connect(server=server, user=username, password=password, database=database)
    cursor = conn.cursor()
    try:
        if nowyPowod == "":
            zapytanie = f"UPDATE Osoba SET stan='obecny', powod='{nowyPowod}' WHERE imie_nazwisko = '{osobaDoZmiany}'"
            cursor.execute(zapytanie)
            conn.commit()
        else:
            zapytanie = f"UPDATE Osoba SET stan='nieobecny', powod='{nowyPowod}' WHERE imie_nazwisko = '{osobaDoZmiany}'"
            cursor.execute(zapytanie)
            conn.commit()
    except Exception as e:
        print(f'Błąd: {e}')
    finally:
        conn.close()

def ZliczanieStanuCalosciowego():
    def createSheet(path):
        xlsxFile = Path(path)
        wbObj = openpyxl.load_workbook(xlsxFile)
        return wbObj
    workbench = createSheet(pathToIRB)
    sheet = workbench.active
    sheet['G4'] = 0; sheet['H4'] = 0; sheet['I4'] = 0; sheet['J4'] = 0; sheet['K4'] = 0; sheet['L4'] = 0; sheet['M4'] = 0; sheet['N4'] = 0
    zliczanie = 0
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=5)):
        for cell in row:
            if cell.value == None:
                sheet['H4'] = int(sheet['H4'].value) + 1
            elif cell.value == 'L4' or cell.value == 'l4':
                sheet['M4'] = int(sheet['M4'].value) + 1
            elif cell.value == 'pj':
                sheet['I4'] = int(sheet['I4'].value) + 1
            elif cell.value == 'ps':
                sheet['J4'] = int(sheet['J4'].value) + 1
            elif cell.value == 'u':
                sheet['K4'] = int(sheet['K4'].value) + 1
            elif cell.value == 'psl':
                sheet['L4'] = int(sheet['L4'].value) + 1
            elif cell.value == 'sl':
                sheet['N4'] = int(sheet['N4'].value) + 1
            zliczanie += 1
    sheet['G4'] = zliczanie
    workbench.save(pathToIRB)

def TworzenieExcelaZBD(numerJednostki):
    server = 'dbserver-projekt-zespolowy-dt.database.windows.net'
    database = 'db'
    username = 'Ladybug'
    password = 'FfqGV3PY'
    conn = pymssql.connect(server=server, user=username, password=password, database=database)
    cursor = conn.cursor()
    query = f"SELECT stopien, imie_nazwisko, stan, powod, kompania_id FROM Osoba"
    cursor.execute(query)
    rows = cursor.fetchall()
    workbook = Workbook()
    sheet = workbook.active
    naglowki = ['Lp.', 'Stopień.', 'Nazwisko i imię', 'Obecny / Nieobecny', 'Powód', '']
    sheet.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = 5)
    sheet['A1'] = str(numerJednostki) + " Kompania"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.append(naglowki)
    zliczanie = 1
    for row in rows:
        if row[4] == numerJednostki:
            sheet.append([zliczanie, row[0], row[1], row[2], row[3]])
            zliczanie += 1
    sheet.merge_cells(start_row = 2, start_column = 7, end_row = 2, end_column = 14)
    sheet['G2'] = "Stan całościowy"
    sheet['G2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['G3'] = 'SE'; sheet['H3'] = 'SF'; sheet['I3'] = 'PJ'; sheet['J3'] = 'PS'; sheet['K3'] = 'U'; sheet['L3'] = 'PSŁ'; sheet['M3'] = 'L4'; sheet['N3'] = 'SŁ'
    sheet['G4'] = 0; sheet['H4'] = 0; sheet['I4'] = 0; sheet['J4'] = 0; sheet['K4'] = 0; sheet['L4'] = 0; sheet['M4'] = 0; sheet['N4'] = 0
    workbook.save(pathToIRB)
    ZliczanieStanuCalosciowego()

def createSheet(path):
    xlsxFile = Path(path)
    wbObj = openpyxl.load_workbook(xlsxFile)
    return wbObj

def PobierzDane(name):
    workbench = createSheet(pathToPrzepustki)
    sheet = workbench.active
    for i, row in enumerate(sheet.iter_rows(min_row=sheet['F3'].value + 5 + sheet['F5'].value + 1, min_col=2, max_col=2)):
        for cell in row:
            if cell.value == name:
                return "Zakaz", "", ""
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet['F3'].value + 2, min_col=2, max_col=2)):
        for cell in row:
            if cell.value == name:
                return "u", sheet.cell(cell.row, cell.column + 1).value.strftime("%d.%m.%Y"), sheet.cell(cell.row, cell.column + 2).value.strftime("%d.%m.%Y")
    for i, row in enumerate(sheet.iter_rows(min_row=sheet['F3'].value + 5, max_row=sheet['F3'].value + sheet['F5'].value + 4, min_col=2, max_col=2)):
        for cell in row:
            if cell.value == name:
                return "pj", sheet.cell(cell.row, cell.column + 1).value.strftime("%d.%m.%Y"), sheet.cell(cell.row, cell.column + 2).value.strftime("%d.%m.%Y")
    return "", "", ""

def JakaPrzepustka(name):
    rodzajPrzepustki, od, do = PobierzDane(name)
    formattedDate = date.today().strftime("%d.%m.%Y")
    actualDay = datetime.today().weekday()
    formattedTime = datetime.now().strftime("%H.%M")
    if rodzajPrzepustki == "Zakaz":
        return "", "", ""
    if (rodzajPrzepustki == "u" and str(formattedTime) >= "12.00" and formattedDate == od) or (rodzajPrzepustki == "u" and (od < formattedDate <= do)):
        return "u", od, do
    elif (rodzajPrzepustki == "pj" and str(formattedTime) >= "15.30" and formattedDate == od) or (rodzajPrzepustki == "pj" and (od < formattedDate <= do)):
        return "pj", od, do
    else:
        if actualDay == 5 and (str(formattedTime) >= "12.00"):
            return "ps", date.today().strftime("%d.%m.%Y"), (date.today() + timedelta(days=1)).strftime("%d.%m.%Y")
        elif actualDay == 6 or ((actualDay == 3 or actualDay == 2 or actualDay == 1 or actualDay == 0) and str(formattedTime) >= "15.30"):
            return "ps", date.today().strftime("%d.%m.%Y"), date.today().strftime("%d.%m.%Y")
        else:
            return "", "", ""

def ZmienWKsiazce(name, przepustka):
    if przepustka != "":
        przepustka, od, do = JakaPrzepustka(name)
        return przepustka
    else:
        return ""

def InfoZeWypisalBudyn(zolnierz, jakaPrzepustka):
    oknoKomunikatu = Tk()

    def close_window():
        oknoKomunikatu.destroy()

    oknoKomunikatu.title("Komunikat")
    x = (oknoKomunikatu.winfo_screenwidth() - 500) // 2
    y = (oknoKomunikatu.winfo_screenheight() - 60) // 2
    oknoKomunikatu.geometry('{}x{}+{}+{}'.format(500, 60, x, y))
    if jakaPrzepustka == "":
        etykieta = Label(oknoKomunikatu, text=zolnierz + " wrócił na pododdział!", justify=CENTER, font=50, fg="black")
    else:
        etykieta = Label(oknoKomunikatu, text="Wypisałem " +zolnierz + " na " +jakaPrzepustka +"!", justify=CENTER, font=50, fg="black")
    etykieta.pack()
    oknoKomunikatu.after(3000, close_window)
    oknoKomunikatu.mainloop()

def InfoZeBrakMozliwosci(zolnierz):
    oknoKomunikatu = Tk()

    def close_window():
        oknoKomunikatu.destroy()

    oknoKomunikatu.title("Komunikat")
    x = (oknoKomunikatu.winfo_screenwidth() - 500) // 2
    y = (oknoKomunikatu.winfo_screenheight() - 60) // 2
    oknoKomunikatu.geometry('{}x{}+{}+{}'.format(500, 60, x, y))
    etykieta = Label(oknoKomunikatu, text="Niestety nie mogę wypisać "+zolnierz, justify=CENTER, font=50, fg="black")
    etykieta.pack()
    oknoKomunikatu.after(3000, close_window)
    oknoKomunikatu.mainloop()

def ZmianaIRB(name):
    workbook = createSheet(pathToIRB)
    sheet = workbook.active
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet['G4'].value + 3, min_col=3, max_col=3)):
        for cell in row:
            if cell.value == name:
                coWpisac = ""
                if sheet.cell(cell.row, cell.column + 2).value is None:
                    coWpisac, od, do = JakaPrzepustka(name)
                    if coWpisac == "":
                        threading.Thread(target=InfoZeBrakMozliwosci, args=(name,)).start()
                        return "!"
                coWpisac = ZmienWKsiazce(name, coWpisac)
                threading.Thread(target=InfoZeWypisalBudyn, args=(name, coWpisac,)).start()
                sheet.cell(cell.row, cell.column + 2).value = coWpisac
                if coWpisac == "":
                    sheet.cell(cell.row, cell.column + 1).value = "obecny"
                else:
                    sheet.cell(cell.row, cell.column + 1).value = "nieobecny"
                EdycjaBazyDanych(sheet.cell(cell.row, cell.column).value, coWpisac)
                workbook.save(pathToIRB)
                return coWpisac

def ObslugaUsera(name):
    jakiEfekt = ZmianaIRB(name)
    if jakiEfekt != "!" and jakiEfekt != "":
        print("Wypisałem " +name+" na " +jakiEfekt +".")
    elif jakiEfekt == "":
        print(name +" wrócił na teren pododdziału.")
    else:
        print("Niestety nie mogę wypisać " +name +".")

def KomunikatKtoNieWrocil(zolnierz):
    with open(pathToNumer, 'r') as numerki:
        for i in numerki.readlines():
            osoba, numer = str(i).split(":")
            if osoba == zolnierz:
                oknoKomunikatu = Tk()
                oknoKomunikatu.title("Komunikat")
                x = (oknoKomunikatu.winfo_screenwidth() - 600) // 2
                y = (oknoKomunikatu.winfo_screenheight() - 60) // 2
                oknoKomunikatu.geometry('{}x{}+{}+{}'.format(600, 60, x, y))
                etykieta = Label(oknoKomunikatu, text="  Żołnierz " + zolnierz + " jeszcze nie wrócił, zostało mu mało czasu!  \n  " +str(numer).replace("\n", "").replace(" ", ""), justify=CENTER, font=30, fg="red")
                etykieta.pack()
                oknoKomunikatu.mainloop()
                break

def SprawdzenieCzyKoniecPrzepustki():
    workbench = createSheet(pathToIRB)
    sheet = workbench.active
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=6)):
        if sheet.cell(i + 3, 5).value == 'ps':
            threading.Thread(target=KomunikatKtoNieWrocil, args=(sheet.cell(i + 3, 3).value,)).start()

def ZwrotKtoSluzba():
    workbench = createSheet(pathToIRB)
    sheet = workbench.active
    zbierzOsoby = []
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet['G4'].value + 3, min_col=5, max_col=5)):
        for cell in row:
            if cell.value == "sl":
                zbierzOsoby.append(sheet.cell(cell.row, cell.column - 2).value)
    return zbierzOsoby

def PrzypomnienieOZmianeSluzby():
    oknoKomunikatu = Tk()
    oknoKomunikatu.title("Komunikat")
    x = (oknoKomunikatu.winfo_screenwidth() - 600) // 2
    y = (oknoKomunikatu.winfo_screenheight() - 60) // 2
    oknoKomunikatu.geometry('{}x{}+{}+{}'.format(600, 60, x, y))
    etykieta = Label(oknoKomunikatu, text="Przypominam, że jest jeszcze wpisana stara służba!", justify=CENTER, font=30, fg="red")
    etykieta.pack()
    oknoKomunikatu.mainloop()

def Archiwizacja():
    formatted_date = datetime.now().strftime("%d.%m.%Y")
    if not os.path.exists(f'{pathToArchive}/{formatted_date}'):
        os.makedirs(f'{pathToArchive}/{formatted_date}')
    for i, czas in enumerate(archiwizacjaDokumentu):
        if datetime.now().strftime("%H.%M.%S") == czas:
            shutil.copy2(pathToIRB, f'{pathToArchive}/{formatted_date}/Imienne_Rozliczenie_Bojowe{i}.xlsx')
            break

def WykrywanieTwarzy():
    camera = cv.VideoCapture(0)
    dur2 = 10
    time.sleep(3)
    dur3 = 2
    dur = 2
    dur4 = 2
    osobyNaSl = []
    while cv.waitKey(1) & 0xFF != ord('q'):
        _, img = camera.read()
        grey = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
        faces = faceClassifier.detectMultiScale(grey, scaleFactor=1.1, minNeighbors=4)
        if time.time() - dur > 4:
            for czas in powiadomienieOStarejSluzbie:
                if (datetime.now() + timedelta(hours=1)).strftime("%d.%m.%Y") == czas:
                    osobyNaSl = ZwrotKtoSluzba()
                    dur = time.time()
                    break
        if (time.time() - dur3 > 4):
            for czas in powiadomienieOKoncuPrzepustki:
                if datetime.now().strftime("%H.%M.%S") == czas:
                    SprawdzenieCzyKoniecPrzepustki()
                    dur3 = time.time()
                    break
        if(time.time() - dur > 4) and osobyNaSl == ZwrotKtoSluzba():
            for czas in powiadomienieOStarejSluzbie:
                if datetime.now().strftime("%H.%M.%S") == czas:
                    threading.Thread(target=PrzypomnienieOZmianeSluzby).start()
                    dur = time.time()
                    break
        if(time.time() - dur4 > 4):
            for czas in archiwizacjaDokumentu:
                if datetime.now().strftime("%H.%M.%S") == czas:
                    threading.Thread(target=Archiwizacja).start()
                    dur4 = time.time()
                    break
        for x, y, w, h in faces:
            faceRegion = grey[y:y + h, x:x + w]
            faceRegion = cv.resize(faceRegion, (220, 220))
            label, trust = lbph.predict(faceRegion)
            if (trust < 53):  # złoty środek jak narazie 53
                try:
                    name = id_names[id_names['id'] == label]['name'].item()
                    cv.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 2)
                    cv.putText(img, name, (x, y + h + 30), cv.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255))
                    if (time.time() - dur2 > 4):
                        ObslugaUsera(str(name))
                        dur2 = time.time()
                except:
                    pass
        cv.imshow('Obraz z kamery', img)
    camera.release()
    cv.destroyAllWindows()

def dodanieZdjecia():
    if os.path.exists('id_names.csv'):
        id_names = pd.read_csv('id_names.csv')
        id_names = id_names[['id', 'name']]
    else:
        id_names = pd.DataFrame(columns=['id', 'name'])
        id_names.to_csv('id_names.csv')

    if not os.path.exists('faces'):
        os.makedirs('faces')

    id = int(input('Podaj id nowej lub edytowalnej osoby: '))
    name = ''

    if id in id_names['id'].values:
        name = id_names[id_names['id'] == id]['name'].item()
        print(f'Dzien dobry {name}!!')
    else:
        name = input('Wpisz swoje imie: ')
        os.makedirs(f'faces/{id}')
        id_names = id_names.append({'id': id, 'name': name}, ignore_index=True)
        id_names.to_csv('id_names.csv')

    camera = cv.VideoCapture(0)
    face_classifier = cv.CascadeClassifier('Classifiers/haarface.xml')

    photos_taken = 0

    while (cv.waitKey(1) & 0xFF != ord('q')):
        _, img = camera.read()
        grey = cv.cvtColor(img, cv.COLOR_BGR2GRAY)
        faces = face_classifier.detectMultiScale(grey, scaleFactor=1.1, minNeighbors=5, minSize=(50, 50))
        for (x, y, w, h) in faces:
            cv.rectangle(img, (x, y), (x + w, y + h), (0, 0, 255), 2)

            face_region = grey[y:y + h, x:x + w]
            if cv.waitKey(1) & 0xFF == ord('s') and np.average(face_region) > 50:
                face_img = cv.resize(face_region, (220, 220))
                img_name = f'face.{id}.{datetime.now().microsecond}.jpeg'
                cv.imwrite(f'faces/{id}/{img_name}', face_img)
                photos_taken += 1
                print(f'Zdjecie nr: {photos_taken}')

        cv.imshow('Face', img)
    print('Rozpoczecie lekcji')
    trenowanie()
    trenowanie()
    print('Koniec lekcji')
    camera.release()
    cv.destroyAllWindows()

def trenowanie():
    id_names = pd.read_csv('id_names.csv')
    id_names = id_names[['id', 'name']]
    lbph = cv.face.LBPHFaceRecognizer_create(threshold=500)
    def create_train():
        faces = []
        labels = []
        for id in os.listdir('faces'):
            path = os.path.join('faces', id)
            try:
                os.listdir(path)
            except:
                continue
            for img in os.listdir(path):
                try:
                    face = cv.imread(os.path.join(path, img))
                    face = cv.cvtColor(face, cv.COLOR_BGR2GRAY)

                    faces.append(face)
                    labels.append(int(id))
                except:
                    pass
        return np.array(faces), np.array(labels)
    faces, labels = create_train()
    lbph.train(faces, labels)
    lbph.save('Classifiers/TrainedLBPH.yml')

def haszuj_sha3(tekst):
    sha3_hasz = hashlib.sha3_256()
    sha3_hasz.update(tekst.encode('utf-8'))
    zhashowany_wynik = sha3_hasz.hexdigest()
    return zhashowany_wynik

class Ui_Logowanie(QtWidgets.QDialog):
    def __init__(self):
        super(Ui_Logowanie, self).__init__()
        self.setWindowIcon(QtGui.QIcon("ikona.png"))
        self.initUI()
        self.ui.pushButton.clicked.connect(self.Zaloguj)
        self.ui.pushButton_2.clicked.connect(self.Wyjdz)

    def initUI(self):
        uic.loadUi('PlikiUi/OknoLogowania.ui', self)
        self.ui = Ui_Dialog()
        self.ui.setupUi(self)

    def Zaloguj(self):
        login = self.ui.lineEdit.text()
        haslo = self.ui.lineEdit_2.text()
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT password FROM users WHERE username = '{haszuj_sha3(login)}'"
        cursor.execute(query)
        rows = cursor.fetchall()
        pobraneHaslo = ''
        for row in rows:
            pobraneHaslo = format(row[0])
            break
        conn.close()
        if pobraneHaslo == haszuj_sha3(haslo):
            letters = ""
            try:
                letters = ''.join(filter(str.isalpha, login))
            except:
                self.showAdditionalWindow("Błędny login!", "Błąd")
            digits = ""
            if letters != 'ODWat' and letters != 'ODPion' and letters != 'Admin' and letters != 'Komendant':
                try:
                    digits = int(''.join(filter(str.isdigit, login)))
                except:
                    self.showAdditionalWindow("Błędny login!", "Błąd")
            if letters == 'Kompania':
                self.close()
                self.showAdditionalWindow(f'Witam {str(digits)} Kompanie', "Witam :)")
                TworzenieExcelaZBD(digits)
                threading.Thread(target=WykrywanieTwarzy).start()
                Ui_Glowne(f'{str(digits)} Kompania').exec_()
            elif letters == 'Batalion':
                self.close()
                self.showAdditionalWindow(f'Witam {str(digits)} Batalion', "Witam :)")
                Ui_Glowne(f'{str(digits)} Batalion').exec_()
            elif letters == 'ODPion':
                if digits == 0:
                    digits = ''
                self.close()
                self.showAdditionalWindow(f'Witam służbę oficera {str(digits)} Pionu', "Witam :)")
                Ui_Glowne(f'{str(digits)} ODPion').exec_()
            elif letters == 'ODWat':
                self.close()
                self.showAdditionalWindow(f'Witam służbę oficera Wat', "Witam :)")
                Ui_Glowne('ODWat').exec_()
            elif letters == 'Admin':
                self.close()
                self.showAdditionalWindow(f'Witam Admina kochanego', "Witam :)")
                Ui_PanelAdministratora().exec_()
            elif letters == 'Komendant':
                self.close()
                self.showAdditionalWindow(f'Witam', "Witam :)")
                Ui_PanelGeneratoraHasel().exec_()
            elif letters == 'Dowodca':
                self.close()
                self.showAdditionalWindow(f'Witam dowódce {str(digits)} Kompani', "Witam :)")
                threading.Thread(target=dodanieZdjecia).start()
                self.showAdditionalWindow(f'Wpisz id nowej lub edytowanej osoby.\nJeżeli osoba jest nowa wpisz jej imię i nazwisko\nNaciśnij s by zrobić nowe zdjęcie\nNaciśnij q w celu zakończenia procesu.', "Zasady!")
            else:
                self.showAdditionalWindow("Błędny login!", "Błąd")
        else:
            self.showAdditionalWindow("Błędne hasło!", "Błąd")

    def Wyjdz(self):
        sys.exit(app.exec_())

    def showAdditionalWindow(self, message_text, title_text):
        self.additional_window = QDialog(self)
        self.setWindowIcon(QtGui.QIcon("ikona.png"))
        self.additional_window.setWindowTitle(title_text)
        self.additional_window.setGeometry(200, 200, 300, 100)
        layout = QVBoxLayout()
        label = QLabel(message_text)
        font = QFont()
        font.setPointSize(12)
        label.setFont(font)
        layout.addWidget(label)
        self.additional_window.setLayout(layout)
        timer = QTimer(self)
        timer.timeout.connect(self.closeAdditionalWindow)
        timer.start(30000)
        self.additional_window.exec_()

    def closeAdditionalWindow(self):
        self.additional_window.accept()

class MyThread(threading.Thread):
    update_signal = pyqtSignal()

    def __init__(self, userName):
        super(MyThread, self).__init__()
        self.userName = userName
        self.stop_thread = threading.Event()

    def run(self):
        while not self.stop_thread.is_set():
            self.ControlIfSignal(self.userName)
            time.sleep(20)

    def stop(self):
        self.stop_thread.set()

    def ControlIfSignal(self, userName):
        name = ''.join(filter(str.isalpha, userName))
        try:
            number = int(''.join(filter(str.isdigit, userName)))
        except:
            number = 1
        if name == 'Batalion':
            rows = self.ZwrocSygnalDlaCiebie(name + str(number))
            if len(rows) != 0:
                for row in rows:
                    WyswietlenieKomunikatu(format(row[0]), format(row[1]))
                    self.DeleteZBazyDanychSygnaly(name + str(number))
        elif name == 'ODPion':
            rows = self.ZwrocSygnalDlaCiebie(name)
            if len(rows) != 0:
                for row in rows:
                    WyswietlenieKomunikatu(format(row[0]), format(row[1]))
                    self.DeleteZBazyDanychSygnaly(name)
        elif name == 'Kompania':
            rows = self.ZwrocSygnalDlaCiebie(name + str(number))
            if len(rows) != 0:
                for row in rows:
                    WyswietlenieKomunikatu(format(row[0]), format(row[1]))
                    self.DeleteZBazyDanychSygnaly(name + str(number))

    def ZwrocSygnalDlaCiebie(self, nazwa):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT rodzaj_alarmu, kto_wprowadzil FROM sygnaly_obsluga WHERE pododzial = '{nazwa}'"
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    def DeleteZBazyDanychSygnaly(self, pododzial):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        try:
            query = f"DELETE FROM sygnaly_obsluga WHERE pododzial='{pododzial}'"
            cursor.execute(query)
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Błąd: {e}")
            conn.rollback()

def WyswietlenieKomunikatu(komunikat, odKogo):
    oknoKomunikatu = Tk()
    oknoKomunikatu.title("Komunikat")
    x = (oknoKomunikatu.winfo_screenwidth() - 600) // 2
    y = (oknoKomunikatu.winfo_screenheight() - 60) // 2
    oknoKomunikatu.geometry('{}x{}+{}+{}'.format(600, 60, x, y))
    etykieta = Label(oknoKomunikatu, text=f"Dostałeś sygnał / komunikat '{komunikat}' od {odKogo}.\nOdebrałeś go o {get_formatted_time()}.", justify=CENTER, font=30, fg="red")
    etykieta.pack()
    oknoKomunikatu.mainloop()

def get_formatted_time():
    current_time = time.localtime()
    formatted_time = time.strftime("%H:%M", current_time)
    return formatted_time

class Ui_Glowne(QtWidgets.QDialog):
    pathToIRB = "Imienne_Rozliczenie_Bojowe2.xlsx"
    pathToKsiazka = "Dokumenty/Ksiazka_Wychodzacych.xlsx"
    pathToPrzepustki = "Dokumenty/Przepustki.xlsx"
    pathToNumer = "Dokumenty/Spis_Numerow.txt"
    pathToSluzba = "Dokumenty/Lista_Sluzb.xlsx"

    def __init__(self, userName):
        super(Ui_Glowne, self).__init__()
        self.setWindowIcon(QtGui.QIcon("ikona.png"))
        self.initUI(userName)
        self.ui.label_12.setText(userName)
        self.ui.pushButton_4.clicked.connect(self.WypiszL4)
        self.ui.pushButton_5.clicked.connect(self.ObejmijZdejmijSluzbe)
        my_thread = MyThread(userName)
        my_thread.start()
        self.ui.pushButton_6.clicked.connect(lambda: self.WymusAtualizacje(userName))
        self.ui.pushButton_2.clicked.connect(lambda: self.WyslijKomunikat(userName, self.ui.lineEdit_2.text()))
        self.ui.pushButton_3.clicked.connect(lambda: self.WyslijKomunikat(userName, self.ui.comboBox_2.currentText()))
        self.ui.pushButton.clicked.connect(lambda: self.Wyjdz(my_thread))

    def initUI(self, userName):
        uic.loadUi('PlikiUi/OknoGlowne.ui', self)
        self.ui = Ui_Dialog3()
        self.ui.setupUi(self)
        name = ''.join(filter(str.isalpha, userName))
        try:
            number = int(''.join(filter(str.isdigit, userName)))
        except:
            number = 0
        if name == 'Kompania':
            self.ui.label_4.setEnabled(False)
            self.ui.label_3.setEnabled(False)
            self.ui.label_2.setEnabled(False)
            self.ui.label.setEnabled(False)
            self.ui.lineEdit_2.setEnabled(False)
            self.ui.comboBox.setEnabled(False)
            self.ui.comboBox_2.setEnabled(False)
            self.ui.pushButton_2.setEnabled(False)
            self.ui.pushButton_3.setEnabled(False)
        else:
            self.ui.label_8.setEnabled(False)
            self.ui.label_5.setEnabled(False)
            self.ui.label_6.setEnabled(False)
            self.ui.label_7.setEnabled(False)
            self.ui.comboBox_3.setEnabled(False)
            self.ui.comboBox_4.setEnabled(False)
            self.ui.lineEdit.setEnabled(False)
            self.ui.pushButton_4.setEnabled(False)
            self.ui.pushButton_5.setEnabled(False)
        self.wypelnijComboBox_2()
        self.wypelnijComboBox_3()
        self.wypelnijComboBox_4()
        self.wypelnijComboBox(name, number)
        self.Aktualizacja()

    def pobierzDokumentIRB(self):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT username FROM users"
        cursor.execute(query)
        rows = cursor.fetchall()
        conn.close()

    def wypelnijComboBox(self, name, number):
        if name == 'ODWat' or name == 'ODPion':
            query = f"SELECT username FROM users"
            conn = pymssql.connect(server=server, user=username, password=password, database=database)
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            conn.close()
            for row in rows:
                for i in range(20):
                    if haszuj_sha3('Kompania' + str(i)) == format(row[0]):
                        self.ui.comboBox.addItem('Kompania' + str(i))
        elif name == 'Batalion':
            query = f"SELECT numer_kompanii FROM Kompania WHERE batalion_id = {number}"
            conn = pymssql.connect(server=server, user=username, password=password, database=database)
            cursor = conn.cursor()
            cursor.execute(query)
            rows = cursor.fetchall()
            conn.close()
            for row in rows:
                for i in range(20):
                    if haszuj_sha3('Kompania' + str(i)) == haszuj_sha3(format(row[0])):
                        self.ui.comboBox.addItem('Kompania' + str(i))

    def wypelnijComboBox_2(self):
        self.ui.comboBox_2.clear()
        self.ui.comboBox_2.addItem("--", 0)
        self.ui.comboBox_2.addItem("Alfa", 1)
        self.ui.comboBox_2.addItem("Brawo", 2)
        self.ui.comboBox_2.addItem("Charli", 3)
        self.ui.comboBox_2.addItem("Delta", 4)
        self.ui.comboBox_2.addItem("Alfa crp", 5)
        self.ui.comboBox_2.addItem("Brawo crp", 6)
        self.ui.comboBox_2.addItem("Charli crp", 7)
        self.ui.comboBox_2.addItem("Delta crp", 8)
        self.ui.comboBox_2.addItem("Jodła 1", 9)
        self.ui.comboBox_2.addItem("Jodła 2", 10)
        self.ui.comboBox_2.addItem("Jodła 3", 11)
        self.ui.comboBox_2.addItem("Wzmocnienie", 12)
        self.ui.comboBox_2.addItem("Termostat", 13)
        self.ui.comboBox_2.addItem("Obsada", 14)
        self.ui.comboBox_2.addItem("Posterunek", 15)
        self.ui.comboBox_2.addItem("Kurier", 16)
        self.ui.comboBox_2.addItem("Strefa", 17)
        self.ui.comboBox_2.addItem("Strefa III", 18)

    def wypelnijComboBox_3(self):
        self.ui.comboBox_3.clear()
        self.ui.comboBox_3.addItem("--", 0)
        workbench = createSheet(pathToIRB)
        sheet = workbench.active
        zliczanie = 1
        for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=6)):
            if sheet.cell(i + 3, 5).value == None and sheet.cell(i + 3, 5).value != "l4":
                self.ui.comboBox_3.addItem(sheet.cell(i + 3, 3).value, zliczanie)
                zliczanie += 1

    def wypelnijComboBox_4(self):
        self.ui.comboBox_4.clear()
        self.ui.comboBox_4.addItem("--", 0)
        workbench = createSheet(pathToSluzba)
        sheet = workbench.active
        wczoraj = date.today() - timedelta(days=1)
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=1)):
            for j, cell in enumerate(row):
                if str(sheet.cell(cell.row, cell.column).value) == str(date.today().strftime("%Y-%m-%d") + " 00:00:00") or str(sheet.cell(cell.row, cell.column).value) == str(wczoraj.strftime("%Y-%m-%d") + " 00:00:00"):
                    self.ui.comboBox_4.addItem(str(sheet.cell(cell.row, cell.column+2).value), str(sheet.cell(cell.row, cell.column+2).value))
                    self.ui.comboBox_4.addItem(str(sheet.cell(cell.row, cell.column + 4).value), str(sheet.cell(cell.row, cell.column + 4).value))

    def ObejmijZdejmijSluzbe(self):
        osobaDoZmiany = self.ui.comboBox_4.currentText()
        if osobaDoZmiany == "--":
            self.ui.label_13.setText("Osoba do zmiany służby musi być wybrana!")
        else:
            workbench = createSheet(pathToIRB)
            sheet = workbench.active
            for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=3, max_col=3)):
                for j, cell in enumerate(row):
                    if osobaDoZmiany == cell.value:
                        if sheet.cell(cell.row, cell.column + 2).value == None:
                            sheet.cell(cell.row, cell.column + 2).value = "sl"
                            sheet.cell(cell.row, cell.column + 1).value = "nieobecny"
                            self.ui.label_13.setText(osobaDoZmiany + " objeła służbę.")
                        elif sheet.cell(cell.row, cell.column + 2).value == "sl":
                            sheet.cell(cell.row, cell.column + 2).value = ""
                            sheet.cell(cell.row, cell.column + 1).value = "obecny"
                            self.ui.label_13.setText(osobaDoZmiany + " zdała służbę.")
                        break
            workbench.save(pathToIRB)

    def WypiszL4(self):
        osobaNaL4 = self.ui.comboBox_3.currentText()
        doKiedy = self.ui.lineEdit.text()
        if len(doKiedy) != 10:
            self.ui.label_13.setText("Data musi mieć długość 10 znaków!")
        elif doKiedy.count('.') != 2:
            self.ui.label_13.setText("Data musi być formatu: dd.mm.YYYY !")
        try:
            datetime.strptime(doKiedy, "%d.%m.%Y")
            self.ui.label_13.setText(osobaNaL4 + " została wpisana na L4!")
            if doKiedy <= date.today().strftime("%d.%m.%Y"):
                self.ui.label_13.setText("Data nie może być dzisiejsza lub przeszła!")
            else:
                self.ZmianaIRB(osobaNaL4, doKiedy)
                self.wypelnijComboBox_3()
                self.ui.lineEdit.setText("")
        except ValueError:
            self.ui.label_13.setText("Źle podana data, przypominam o formacie: dd.mm.YYYY !")

    def ZmianaIRB(self, name, doKiedy):
        workbook = createSheet(pathToIRB)
        sheet = workbook.active
        for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet['G4'].value + 3, min_col=3, max_col=3)):
            for cell in row:
                if cell.value == name:
                    sheet.cell(cell.row, cell.column + 2).value = "l4"
                    sheet.cell(cell.row, cell.column + 1).value = "nieobecny"
                    self.ZmienWKsiazce(name, "", sheet.cell(cell.row, cell.column - 1).value, doKiedy)
                    self.ZmienWKsiazce(name, "l4", sheet.cell(cell.row, cell.column - 1).value, doKiedy)
                    workbook.save(pathToIRB)

    def Aktualizacja(self):
        self.wypelnijComboBox_3()
        self.wypelnijComboBox_4()
        self.ui.label_13.setText("")
        self.ui.textBrowser.setText("")
        self.ui.textBrowser_2.setText("")
        self.ui.textBrowser_3.setText("")

    def WymusAtualizacje(self, userName):
        name = ''.join(filter(str.isalpha, userName))
        try:
            number = int(''.join(filter(str.isdigit, userName)))
        except:
            number = 1
        if name == 'Kompania':
            numerKompania = number
        else:
            tekstZComboBox = self.ui.comboBox.currentText()
            numerKompania = int(''.join(filter(str.isdigit, tekstZComboBox)))
        TworzenieExcelaZBD(numerKompania)
        self.ui.textBrowser.setText("")
        self.ui.textBrowser_2.setText("")
        self.ui.textBrowser_3.setText("")
        self.ui.label_13.setText("")
        workbench = createSheet(pathToIRB)
        sheet = workbench.active
        formatted_line = ""
        for i, row in enumerate(sheet.iter_rows(min_row=2, max_row=sheet['G4'].value + 3, min_col=1, max_col=5)):
            linia = ""
            for j, cell in enumerate(row):
                if j != 3:
                    text = cell.value
                    if cell.value == None:
                        text = "  "
                    if j == 0:
                        spaces = "&nbsp;" * (5 - len(str(text)))
                    elif j == 1:
                        spaces = "&nbsp;" * (15 - len(str(text)))
                    else:
                        spaces = "&nbsp;" * (20 - len(str(text)))
                    linia += f"{text}{spaces}"
            if i == 0:
                formatted_line += f"<pre><b>{linia}</b></pre>"
            else:
                formatted_line += f"<pre>{linia}</pre>"
        self.ui.textBrowser.insertHtml(formatted_line)
        workbench = createSheet(pathToIRB)
        sheet = workbench.active
        formatted_line = f'<pre><b><span style="font-size: 12pt;">SE    SF    PJ    PS    U    PSŁ    L4    SŁ</span></b></pre>'
        ileSF = 0; ilePJ = 0; ilePS = 0; ileU = 0; ilePSL = 0; ileL4 = 0; ileSl = 0
        for k in range(3, sheet.max_row + 1):
            if sheet['E' + str(k)].value == None:
                ileSF += 1
            elif sheet['E' + str(k)].value == "pj":
                ilePJ += 1
            elif sheet['E' + str(k)].value == "ps":
                ilePS += 1
            elif sheet['E' + str(k)].value == "u":
                ileU += 1
            elif sheet['E' + str(k)].value == "psl":
                ilePSL += 1
            elif sheet['E' + str(k)].value == "l4" or sheet['E' + str(k)].value == "L4":
                ileL4 += 1
            elif sheet['E' + str(k)].value == "sl":
                ileSl += 1
        spaces = "&nbsp;"
        linia = f"{str(ileSF + ilePJ + ilePS + ileU + ilePSL + ileL4 + ileSl)}{spaces * (6 - len(str(ileSF + ilePJ + ilePS + ileU + ilePSL + ileL4 + ileSl)))}" \
                f"{str(ileSF)}{spaces * (6 - len(str(ileSF)))}{str(ilePJ)}{spaces * (6 - len(str(ilePJ)))}" \
                f"{str(ilePS)}{spaces * (6 - len(str(ilePS)))}" \
                f"{str(ileU)}{spaces * (6 - len(str(ileU)))}{str(ilePSL)}{spaces * (6 - len(str(ilePSL)))}" \
                f"{str(ileL4)}{spaces * (6 - len(str(ileL4)))}{str(ileSl)}{spaces * (6 - len(str(ileSl)))}"
        formatted_line += f'<pre><b><span style="font-size: 12pt;">{linia}</span></b></pre>'
        self.ui.textBrowser_3.insertHtml(formatted_line)

    def WyslijKomunikat(self, userName, trescKomunikatu):
        if trescKomunikatu == "" or trescKomunikatu == "--":
            return
        name = ''.join(filter(str.isalpha, userName))
        try:
            number = int(''.join(filter(str.isdigit, userName)))
        except:
            number = 1
        if name == 'ODWat':
            pododzial = 'ODPion'
            self.InsertDoBazyDanychSygnaly(trescKomunikatu, pododzial, name)
        elif name == 'ODPion':
            rows = self.pobierzTwojeBataliony()
            for row in rows:
                self.InsertDoBazyDanychSygnaly(trescKomunikatu, format(row[1]), name)
        if name == 'Batalion':
            rows = self.pobierzTwojeKompanie(number)
            for row in rows:
                self.InsertDoBazyDanychSygnaly(trescKomunikatu, format(row[0]), name)
        self.ui.lineEdit_2.setText("")
        self.ui.comboBox_2.setCurrentText("--")

    def pobierzTwojeBataliony(self):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT * FROM Batalion"
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    def pobierzTwojeKompanie(self, number):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT numer_kompanii FROM Kompania WHERE batalion_id = '{number}'"
        cursor.execute(query)
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return rows

    def DeleteZBazyDanychSygnaly(self, pododzial):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        try:
            query = f"DELETE FROM sygnaly_obsluga WHERE pododzial='{pododzial}'"
            cursor.execute(query)
            conn.commit()
            cursor.close()
            conn.close()
        except Exception as e:
            print(f"Błąd: {e}")
            conn.rollback()

    def InsertDoBazyDanychSygnaly(self, rodzaj_alarmu, pododzial, kto_wprowadzil):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT * FROM sygnaly_obsluga"
        cursor.execute(query)
        rows = cursor.fetchall()
        zliczanie = 1
        for row in rows:
            zliczanie = int(format(row[0])) + 1
        nowe_dane = {
            'id': zliczanie,
            'rodzaj_alarmu': rodzaj_alarmu,
            'pododzial': pododzial,
            'kto_wprowadzil': kto_wprowadzil,
        }
        query = f"INSERT INTO sygnaly_obsluga ({', '.join(nowe_dane.keys())}) VALUES ({', '.join(['%s'] * len(nowe_dane))})"
        values = tuple(nowe_dane.values())
        cursor.execute(query, values)
        conn.commit()
        cursor.close()
        conn.close()

    def Wyjdz(self, my_thread):
        my_thread.stop()
        try:
            os.remove(pathToIRB)
        except Exception as e:
            print(f"Wystąpił błąd podczas usuwania pliku: {e}")
        self.close()

class Ui_PanelAdministratora(QtWidgets.QDialog):
    def __init__(self):
        super(Ui_PanelAdministratora, self).__init__()
        self.setWindowIcon(QtGui.QIcon("ikona.png"))
        self.initUI()
        self.cieniowanie()
        self.wypelnijComboBoxy()

    def initUI(self):
        uic.loadUi('PlikiUi/OknoAdministratora.ui', self)
        self.ui = Ui_Dialog4()
        self.ui.setupUi(self)
        self.ui.button_group = QButtonGroup()
        self.ui.button_group.addButton(self.ui.radioButton_4)
        self.ui.button_group.addButton(self.ui.radioButton_5)
        self.ui.button_group.addButton(self.ui.radioButton_10)
        self.ui.button_group.addButton(self.ui.radioButton_7)
        self.ui.button_group.addButton(self.ui.radioButton_6)
        self.ui.button_group.addButton(self.ui.radioButton_8)
        self.ui.button_group.addButton(self.ui.radioButton_9)
        self.ui.button_group2 = QButtonGroup()
        self.ui.button_group2.addButton(self.ui.radioButton)
        self.ui.button_group2.addButton(self.ui.radioButton_2)
        self.ui.button_group2.addButton(self.ui.radioButton_3)
        self.ui.radioButton_4.clicked.connect(self.cieniowanie)
        self.ui.radioButton_5.clicked.connect(self.cieniowanie)
        self.ui.radioButton_6.clicked.connect(self.cieniowanie)
        self.ui.radioButton_7.clicked.connect(self.cieniowanie)
        self.ui.radioButton_8.clicked.connect(self.cieniowanie)
        self.ui.radioButton_9.clicked.connect(self.cieniowanie)
        self.ui.radioButton_10.clicked.connect(self.cieniowanie)
        self.ui.pushButton.clicked.connect(self.wyswietlListe)
        self.ui.pushButton_2.clicked.connect(self.wykonaj)

    def cieniowanie(self):
        self.wycieniujWszystkie()
        if self.ui.radioButton_4.isChecked():
            #Dodaj osobe
            self.ui.comboBox.setEnabled(True)
            self.ui.comboBox_2.setEnabled(True)
            self.ui.lineEdit.setEnabled(True)
            self.ui.label_2.setEnabled(True)
            self.ui.label_3.setEnabled(True)
            self.ui.label_4.setEnabled(True)
        elif self.ui.radioButton_5.isChecked():
            #Usuń osobę
            self.ui.comboBox_3.setEnabled(True)
            self.ui.label_5.setEnabled(True)
        elif self.ui.radioButton_10.isChecked():
            #Edytuj osobę
            self.ui.comboBox_4.setEnabled(True)
            self.ui.comboBox_5.setEnabled(True)
            self.ui.comboBox_6.setEnabled(True)
            self.ui.lineEdit_2.setEnabled(True)
            self.ui.label_10.setEnabled(True)
            self.ui.label_7.setEnabled(True)
            self.ui.label_8.setEnabled(True)
            self.ui.label_9.setEnabled(True)
        elif self.ui.radioButton_7.isChecked():
            #dodaj kompanię
            self.ui.lineEdit_3.setEnabled(True)
            self.ui.comboBox_7.setEnabled(True)
            self.ui.label_6.setEnabled(True)
            self.ui.label_11.setEnabled(True)
        elif self.ui.radioButton_6.isChecked():
            #usuń kompanię
            self.ui.comboBox_8.setEnabled(True)
            self.ui.label_12.setEnabled(True)
        elif self.ui.radioButton_8.isChecked():
            #dodaj batalion
            self.ui.lineEdit_4.setEnabled(True)
            self.ui.label_13.setEnabled(True)
        elif self.ui.radioButton_9.isChecked():
            #usuń batalion
            self.ui.comboBox_9.setEnabled(True)
            self.ui.label_14.setEnabled(True)

    def wycieniujWszystkie(self):
        self.ui.comboBox.setEnabled(False)
        self.ui.comboBox_2.setEnabled(False)
        self.ui.lineEdit.setEnabled(False)
        self.ui.comboBox_3.setEnabled(False)
        self.ui.comboBox_4.setEnabled(False)
        self.ui.comboBox_5.setEnabled(False)
        self.ui.comboBox_6.setEnabled(False)
        self.ui.lineEdit_2.setEnabled(False)
        self.ui.lineEdit_3.setEnabled(False)
        self.ui.comboBox_7.setEnabled(False)
        self.ui.comboBox_8.setEnabled(False)
        self.ui.lineEdit_4.setEnabled(False)
        self.ui.comboBox_9.setEnabled(False)
        self.ui.label_2.setEnabled(False)
        self.ui.label_3.setEnabled(False)
        self.ui.label_4.setEnabled(False)
        self.ui.label_5.setEnabled(False)
        self.ui.label_6.setEnabled(False)
        self.ui.label_7.setEnabled(False)
        self.ui.label_8.setEnabled(False)
        self.ui.label_9.setEnabled(False)
        self.ui.label_10.setEnabled(False)
        self.ui.label_11.setEnabled(False)
        self.ui.label_12.setEnabled(False)
        self.ui.label_13.setEnabled(False)
        self.ui.label_14.setEnabled(False)

    def wypelnijComboBoxy(self):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT id FROM Osoba"
        cursor.execute(query)
        rows = cursor.fetchall()
        self.ui.comboBox.clear()
        self.ui.comboBox.addItem("--", 0)
        self.ui.comboBox.addItem("szer.", 1)
        self.ui.comboBox.addItem("st.szer.", 2)
        self.ui.comboBox.addItem("st.szer.spec.", 3)
        self.ui.comboBox.addItem("kpr.", 4)
        self.ui.comboBox.addItem("st.kpr.", 5)
        self.ui.comboBox.addItem("plut.", 6)
        self.ui.comboBox.addItem("sierż.", 7)
        self.ui.comboBox.addItem("st.sierż.", 8)
        self.ui.comboBox.addItem("mł.chor.", 9)
        self.ui.comboBox.addItem("chor.", 10)
        self.ui.comboBox.addItem("st.chor.", 11)
        self.ui.comboBox.addItem("st.chor.sztab.", 12)
        self.ui.comboBox.addItem("ppor.", 13)
        self.ui.comboBox.addItem("por.", 14)
        self.ui.comboBox.addItem("kpt.", 15)
        self.ui.comboBox.addItem("mjr.", 16)
        self.ui.comboBox.addItem("ppłk.", 17)
        self.ui.comboBox.addItem("płk.", 18)
        self.ui.comboBox.addItem("szer.pchor.", 19)
        self.ui.comboBox.addItem("st.szer.pchor.", 20)
        self.ui.comboBox.addItem("st.szer.spec.pchor.", 21)
        self.ui.comboBox.addItem("kpr.pchor.", 22)
        self.ui.comboBox.addItem("st.kpr.pchor.", 23)
        self.ui.comboBox.addItem("plut.pchor.", 24)
        self.ui.comboBox.addItem("sierż.pchor.", 25)
        self.ui.comboBox_5.addItem("--", 0)
        self.ui.comboBox_5.addItem("szer.", 1)
        self.ui.comboBox_5.addItem("st.szer.", 2)
        self.ui.comboBox_5.addItem("st.szer.spec.", 3)
        self.ui.comboBox_5.addItem("kpr.", 4)
        self.ui.comboBox_5.addItem("st.kpr.", 5)
        self.ui.comboBox_5.addItem("plut.", 6)
        self.ui.comboBox_5.addItem("sierż.", 7)
        self.ui.comboBox_5.addItem("st.sierż.", 8)
        self.ui.comboBox_5.addItem("mł.chor.", 9)
        self.ui.comboBox_5.addItem("chor.", 10)
        self.ui.comboBox_5.addItem("st.chor.", 11)
        self.ui.comboBox_5.addItem("st.chor.sztab.", 12)
        self.ui.comboBox_5.addItem("ppor.", 13)
        self.ui.comboBox_5.addItem("por.", 14)
        self.ui.comboBox_5.addItem("kpt.", 15)
        self.ui.comboBox_5.addItem("mjr.", 16)
        self.ui.comboBox_5.addItem("ppłk.", 17)
        self.ui.comboBox_5.addItem("płk.", 18)
        self.ui.comboBox_5.addItem("szer.pchor.", 19)
        self.ui.comboBox_5.addItem("st.szer.pchor.", 20)
        self.ui.comboBox_5.addItem("st.szer.spec.pchor.", 21)
        self.ui.comboBox_5.addItem("kpr.pchor.", 22)
        self.ui.comboBox_5.addItem("st.kpr.pchor.", 23)
        self.ui.comboBox_5.addItem("plut.pchor.", 24)
        self.ui.comboBox_5.addItem("sierż.pchor.", 25)
        self.ui.comboBox_3.clear()
        self.ui.comboBox_4.clear()
        for row in rows:
            self.ui.comboBox_3.addItem(format(row[0]), int(format(row[0])) - 1)
            self.ui.comboBox_4.addItem(format(row[0]), int(format(row[0])) - 1)
        query = f"SELECT * FROM Batalion"
        cursor.execute(query)
        rows = cursor.fetchall()
        self.ui.comboBox_7.clear()
        self.ui.comboBox_9.clear()
        for row in rows:
            self.ui.comboBox_7.addItem(format(row[1]), int(format(row[0])) - 1)
            self.ui.comboBox_9.addItem(format(row[0]), int(format(row[0])) - 1)
        query = f"SELECT id, numer_kompanii FROM Kompania"
        cursor.execute(query)
        rows = cursor.fetchall()
        self.ui.comboBox_2.clear()
        self.ui.comboBox_6.clear()
        self.ui.comboBox_8.clear()
        for row in rows:
            self.ui.comboBox_2.addItem(format(row[1]), int(format(row[0])) - 1)
            self.ui.comboBox_6.addItem(format(row[1]), int(format(row[0])) - 1)
            self.ui.comboBox_8.addItem(format(row[0]), int(format(row[0])) - 1)

    def wyswietlListe(self):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        self.ui.textBrowser.clear()
        if self.ui.radioButton.isChecked():
            #Lista osób
            query = f"SELECT * FROM Osoba"
            cursor.execute(query)
            rows = cursor.fetchall()
            linia = f"id   stopień   imie_i_nazwisko     powód   id_kompani\n"
            self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
            linia = f"-----------------------------------------------------\n"
            self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
            for row in rows:
                linia = ""
                spaces = "&nbsp;" * (5 - len(format(row[0])))
                linia += f"{format(row[0])}{spaces}"
                spaces = "&nbsp;" * (10 - len(format(row[1])))
                linia += f"{format(row[1])}{spaces}"
                spaces = "&nbsp;" * (20 - len(format(row[2])))
                linia += f"{format(row[2])}{spaces}"
                spaces = "&nbsp;" * (8 - len(format(row[4])))
                linia += f"{format(row[4])}{spaces}"
                spaces = "&nbsp;" * (5 - len(format(row[5])))
                linia += f"{format(row[5])}{spaces}\n"
                self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
        elif self.ui.radioButton_2.isChecked():
            #Lista kompanii
            query = f"SELECT * FROM Kompania"
            cursor.execute(query)
            rows = cursor.fetchall()
            linia = f"id   nazwa_kompani  id_batalionu\n"
            self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
            linia = f"-----------------------------------------------------\n"
            self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
            for row in rows:
                linia = ""
                spaces = "&nbsp;" * (5 - len(format(row[0])))
                linia += f"{format(row[0])}{spaces}"
                spaces = "&nbsp;" * (15 - len(format(row[1])))
                linia += f"{format(row[1])}{spaces}"
                spaces = "&nbsp;" * (5 - len(format(row[2])))
                linia += f"{format(row[2])}{spaces}\n"
                self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
        elif self.ui.radioButton_3.isChecked():
            #Lista batalionów
            query = f"SELECT * FROM Batalion"
            cursor.execute(query)
            rows = cursor.fetchall()
            linia = f"id   nazwa_batalionu\n"
            self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
            linia = f"-----------------------------------------------------\n"
            self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')
            for row in rows:
                linia = ""
                spaces = "&nbsp;" * (5 - len(format(row[0])))
                linia += f"{format(row[0])}{spaces}"
                spaces = "&nbsp;" * (15 - len(format(row[1])))
                linia += f"{format(row[1])}{spaces}\n"
                self.ui.textBrowser.insertHtml(f'<pre><b>{linia}</b></pre>')

    def wykonaj(self):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        if self.ui.radioButton_4.isChecked():
            # Dodaj osobe
            query = f"SELECT * FROM Kompania"
            cursor.execute(query)
            rows = cursor.fetchall()
            nowe_dane = {
                'id': self.ui.comboBox_3.count() + 1,
                'stopien': self.ui.comboBox.currentText(),
                'imie_nazwisko': self.ui.lineEdit.text(),
                'stan': 'obecny',
                'powod': '',
                'kompania_id': ''
            }
            for row in rows:
                if format(row[1]) == self.ui.comboBox_2.currentText():
                    nowe_dane['kompania_id'] = format(row[0])
                    break
            query = f"INSERT INTO Osoba ({', '.join(nowe_dane.keys())}) VALUES ({', '.join(['%s'] * len(nowe_dane))})"
            values = tuple(nowe_dane.values())
            cursor.execute(query, values)
            conn.commit()
            cursor.close()
            conn.close()
            WyswietlenieKomunikatu2(f"Dodałem Osobę o podanych wartościach:\n {nowe_dane['id']} {nowe_dane['stopien']} {nowe_dane['imie_nazwisko']} {nowe_dane['kompania_id']}")
        elif self.ui.radioButton_5.isChecked():
            try:
                query = f"DELETE FROM Osoba WHERE id={self.ui.comboBox_3.currentText()}"
                cursor.execute(query)
                conn.commit()
                WyswietlenieKomunikatu2(f"Usunąłem Osobę o id {self.ui.comboBox_3.currentText()}")
            except Exception as e:
                print(f"Błąd: {e}")
                conn.rollback()
            finally:
                cursor.close()
                conn.close()
        elif self.ui.radioButton_10.isChecked():
            query = f"SELECT * FROM Kompania"
            cursor.execute(query)
            rows = cursor.fetchall()
            noweID = 1
            for row in rows:
                if format(row[1]) == self.ui.comboBox_6.currentText():
                    noweID = format(row[0])
                    break
            try:
                zapytanie = f"UPDATE Osoba SET stopien='{self.ui.comboBox_5.currentText()}', kompania_id={noweID}, imie_nazwisko='{self.ui.lineEdit_2.text()}' WHERE id={self.ui.comboBox_4.currentText()}"
                cursor.execute(zapytanie)
                conn.commit()
            except Exception as e:
                print(f'Błąd: {e}')
            finally:
                conn.close()
            WyswietlenieKomunikatu2(f"Edytowałem Osobe o id {self.ui.comboBox_4.currentText()}")
        elif self.ui.radioButton_7.isChecked():
            query = f"SELECT * FROM Batalion"
            cursor.execute(query)
            rows = cursor.fetchall()
            nowe_dane = {
                'id': self.ui.comboBox_8.count() + 1,
                'numer_kompanii': self.ui.lineEdit_3.text(),
                'batalion_id': ''
            }
            for row in rows:
                if format(row[1]) == self.ui.comboBox_7.currentText():
                    nowe_dane['batalion_id'] = format(row[0])
                    break
            query = f"INSERT INTO Kompania ({', '.join(nowe_dane.keys())}) VALUES ({', '.join(['%s'] * len(nowe_dane))})"
            values = tuple(nowe_dane.values())
            cursor.execute(query, values)
            conn.commit()
            cursor.close()
            conn.close()
            WyswietlenieKomunikatu2(f"Dodałem Kompanię o podanych wartościach:\n {nowe_dane['id']} {nowe_dane['numer_kompanii']} {nowe_dane['batalion_id']}")
        elif self.ui.radioButton_6.isChecked():
            try:
                query = f"DELETE FROM Kompania WHERE id={self.ui.comboBox_8.currentText()}"
                cursor.execute(query)
                conn.commit()
                WyswietlenieKomunikatu2(f"Usunąłem Kompanię o id {self.ui.comboBox_8.currentText()}")
            except Exception as e:
                print(f"Błąd: {e}")
                conn.rollback()
            finally:
                cursor.close()
                conn.close()
        elif self.ui.radioButton_8.isChecked():
            query = f"INSERT INTO Batalion (id, numer_batalionu) VALUES {self.ui.comboBox_9.count() + 1, self.ui.lineEdit_4.text()}"
            cursor.execute(query)
            conn.commit()
            cursor.close()
            conn.close()
            WyswietlenieKomunikatu2(f"Dodałem Batalion o podanych wartościach:\n {self.ui.comboBox_9.count() + 1}, {self.ui.lineEdit_4.text()}")
        elif self.ui.radioButton_9.isChecked():
            try:
                query = f"DELETE FROM Batalion WHERE id={self.ui.comboBox_9.currentText()}"
                cursor.execute(query)
                conn.commit()
                WyswietlenieKomunikatu2(f"Usunąłem Batalion o id {self.ui.comboBox_9.currentText()}")
            except Exception as e:
                print(f"Błąd: {e}")
                conn.rollback()
            finally:
                cursor.close()
                conn.close()
        self.wypelnijComboBoxy()

class Ui_PanelGeneratoraHasel(QtWidgets.QDialog):
    def __init__(self):
        super(Ui_PanelGeneratoraHasel, self).__init__()
        self.setWindowIcon(QtGui.QIcon("ikona.png"))
        self.initUI()

    def initUI(self):
        uic.loadUi('PlikiUi/OknoGeneratoraHasel.ui', self)
        self.ui = Ui_Dialog5()
        self.ui.setupUi(self)
        self.ui.pushButton.clicked.connect(self.pobranieLoginow)

    def closeEvent(self, event):
        self.remove_file()

    def remove_file(self):
        file_path = "PlikiPomocnicze/loginy_hasla.pdf"
        try:
            if os.path.exists(file_path):
                os.remove(file_path)
                print(f"Usunięto plik: {file_path}")
            else:
                print(f"Plik nie istnieje: {file_path}")
        except Exception as e:
            QMessageBox.warning(self, 'Błąd', f'Nie udało się usunąć pliku: {e}')

    def pobranieLoginow(self):
        loginy = []
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        query = f"SELECT numer_batalionu FROM Batalion"
        cursor.execute(query)
        rows = cursor.fetchall()
        for row in rows:
            loginy.append(format(row[0]))
        query = f"SELECT numer_kompanii FROM Kompania"
        cursor.execute(query)
        rows = cursor.fetchall()
        for row in rows:
            loginy.append(format(row[0]))
        loginy.append('ODPion')
        loginy.append('ODWat')
        hasla = generowanieHasel(loginy)
        create_pdf("PlikiPomocnicze/loginy_hasla.pdf", loginy, hasla)
        self.updateBazyDanych(loginy, hasla)

    def updateBazyDanych(self, loginy, hasla):
        conn = pymssql.connect(server=server, user=username, password=password, database=database)
        cursor = conn.cursor()
        for i in range(len(loginy)):
            try:
                update_query = f"UPDATE users SET password = %s WHERE username = %s"
                cursor.execute(update_query, (haszuj_sha3(hasla[i]), haszuj_sha3(loginy[i])))
                conn.commit()
            except Exception as e:
                print(f"Błąd aktualizacji danych: {e}")
        conn.close()

def create_pdf(file_path, logins, passwords):
    logins2, passwords2 = insert_empty_elements(logins, passwords)
    pdf_doc = SimpleDocTemplate(file_path, pagesize=letter)
    data = [["Login", "Haslo"]]
    for login, password in zip(logins2, passwords2):
        data.append([login, password])
    table = Table(data, colWidths=[120, 120])
    style = TableStyle([('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('BACKGROUND', (0, 1), (-1, -1), colors.beige)])
    table.setStyle(style)
    for i in range(1, len(data)):
        table.setStyle([('AFTER', (0, i), (-1, i), 12)])
    story = [table]
    story.append(Spacer(1, 12))
    pdf_doc.build(story)

def WyswietlenieKomunikatu2(tekst):
    oknoKomunikatu = Tk()
    oknoKomunikatu.title("Komunikat")
    x = (oknoKomunikatu.winfo_screenwidth() - 600) // 2
    y = (oknoKomunikatu.winfo_screenheight() - 60) // 2
    oknoKomunikatu.geometry('{}x{}+{}+{}'.format(600, 60, x, y))
    etykieta = Label(oknoKomunikatu, text=f"{tekst}", justify=CENTER, font=30, fg="red")
    etykieta.pack()
    oknoKomunikatu.mainloop()
def generowanieHasel(loginy):
    return [generuj_haslo(16) for _ in range(len(loginy))]
def generuj_haslo(dlugosc):
    znaki = string.ascii_letters + string.digits + string.punctuation
    return ''.join(random.choice(znaki) for _ in range(dlugosc))
def haszuj_sha3(tekst):
    sha3_hasz = hashlib.sha3_256()
    sha3_hasz.update(tekst.encode('utf-8'))
    zhashowany_wynik = sha3_hasz.hexdigest()
    return zhashowany_wynik
def insert_empty_elements(logins, passwords):
    new_logins = []
    new_passwords = []
    for login, password in zip(logins, passwords):
        new_logins.append(login)
        new_passwords.append(password)
        new_logins.append("------------------------------")
        new_passwords.append("------------------------------")
    if new_logins:
        new_logins.pop()
    if new_passwords:
        new_passwords.pop()
    return new_logins, new_passwords

app = QApplication(sys.argv)
ex = Ui_Logowanie()
ex.show()
sys.exit(app.exec_())