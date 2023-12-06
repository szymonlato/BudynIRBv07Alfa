import pymssql
from openpyxl import Workbook
from openpyxl.styles import Alignment
import openpyxl
from pathlib import Path

pathToIRB = "Imienne_Rozliczenie_Bojowe2.xlsx"

def ZliczanieStanuCalosciowego():
    def createSheet(path):
        xlsxFile = Path(path)
        wbObj = openpyxl.load_workbook(xlsxFile)
        return wbObj

    workbench = createSheet(pathToIRB)
    sheet = workbench.active

    sheet['G4'] = 0; sheet['H4'] = 0; sheet['I4'] = 0; sheet['J4'] = 0; sheet['K4'] = 0; sheet['L4'] = 0; sheet['M4'] = 0; sheet['N4'] = 0

    zliczanie = 0
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=5)):
        for cell in row:
            if cell.value == None:
                sheet['H4'] = int(sheet['H4'].value) + 1
            elif cell.value == 'L4':
                sheet['M4'] = int(sheet['M4'].value) + 1
            elif cell.value == 'pj':
                sheet['I4'] = int(sheet['I4'].value) + 1
            elif cell.value == 'ps':
                sheet['J4'] = int(sheet['J4'].value) + 1
            elif cell.value == 'u':
                sheet['K4'] = int(sheet['K4'].value) + 1
            elif cell.value == 'psl':
                sheet['L4'] = int(sheet['L4'].value) + 1
            elif cell.value == 'sl':
                sheet['N4'] = int(sheet['N4'].value) + 1
            zliczanie += 1

    sheet['G4'] = zliczanie
    workbench.save(pathToIRB)

def TworzenieExcelaZBD(numerJednostki):
    server = 'dbserver-projekt-zespolowy-dt.database.windows.net'
    database = 'db'
    username = 'Ladybug'
    password = 'FfqGV3PY'

    conn = pymssql.connect(server=server, user=username, password=password, database=database)
    cursor = conn.cursor()

    query = f"SELECT stopien, imie_nazwisko, stan, powod, kompania_id FROM Osoba"
    cursor.execute(query)
    rows = cursor.fetchall()

    workbook = Workbook()
    sheet = workbook.active

    naglowki = ['Lp.', 'Stopień.', 'Nazwisko i imię', 'Obecny / Nieobecny', 'Powód', '']
    sheet.merge_cells(start_row = 1, start_column = 1, end_row = 1, end_column = 5)
    sheet['A1'] = str(numerJednostki) + " Kompania"
    sheet['A1'].alignment = Alignment(horizontal='center', vertical='center')
    sheet.append(naglowki)
    zliczanie = 1
    for row in rows:
        if row[4] == numerJednostki:
            sheet.append([zliczanie, row[0], row[1], row[2], row[3]])
            zliczanie += 1

    sheet.merge_cells(start_row = 2, start_column = 7, end_row = 2, end_column = 14)
    sheet['G2'] = "Stan całościowy"
    sheet['G2'].alignment = Alignment(horizontal='center', vertical='center')
    sheet['G3'] = 'SE'; sheet['H3'] = 'SF'; sheet['I3'] = 'PJ'; sheet['J3'] = 'PS'; sheet['K3'] = 'U'; sheet['L3'] = 'PSŁ'; sheet['M3'] = 'L4'; sheet['N3'] = 'SŁ'

    sheet['G4'] = 0; sheet['H4'] = 0; sheet['I4'] = 0; sheet['J4'] = 0; sheet['K4'] = 0; sheet['L4'] = 0; sheet['M4'] = 0; sheet['N4'] = 0

    workbook.save(pathToIRB)
    ZliczanieStanuCalosciowego()

def EdycjaBazyDanych(osobaDoZmiany, nowyPowod):
    server = 'dbserver-projekt-zespolowy-dt.database.windows.net'
    database = 'db'
    username = 'Ladybug'
    password = 'FfqGV3PY'

    conn = pymssql.connect(server=server, user=username, password=password, database=database)
    cursor = conn.cursor()
    try:
        if nowyPowod == "":
            zapytanie = f"UPDATE Osoba SET stan='obecny', powod='{nowyPowod}' WHERE imie_nazwisko = '{osobaDoZmiany}'"
            cursor.execute(zapytanie)
            conn.commit()
        else:
            zapytanie = f"UPDATE Osoba SET stan='nieobecny', powod='{nowyPowod}' WHERE imie_nazwisko = '{osobaDoZmiany}'"
            cursor.execute(zapytanie)
            conn.commit()
    except Exception as e:
        print(f'Błąd: {e}')

    finally:
        conn.close()


#Osoba3_9
#Osoba3_6

TworzenieExcelaZBD(6)
#EdycjaBazyDanych("Osoba3_9", "")
#EdycjaBazyDanych("Osoba3_6", "")





