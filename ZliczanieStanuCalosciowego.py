import openpyxl
from pathlib import Path

def createSheet(path):
    xlsxFile = Path(path)
    wbObj = openpyxl.load_workbook(xlsxFile)
    return wbObj

workbench = createSheet("Imienne_Rozliczenie_Bojowe2.xlsx")
sheet = workbench.active

sheet['G4'] = 0; sheet['H4'] = 0; sheet['I4'] = 0; sheet['J4'] = 0; sheet['K4'] = 0; sheet['L4'] = 0; sheet['M4'] = 0; sheet['N4'] = 0

zliczanie = 0
for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=5)):
    for cell in row:
        if cell.value == None:
            sheet['H4'] = int(sheet['H4'].value) + 1
        elif cell.value == 'L4':
            sheet['M4'] = int(sheet['M4'].value) + 1
        elif cell.value == 'pj':
            sheet['I4'] = int(sheet['I4'].value) + 1
        elif cell.value == 'ps':
            sheet['J4'] = int(sheet['J4'].value) + 1
        elif cell.value == 'u':
            sheet['K4'] = int(sheet['K4'].value) + 1
        elif cell.value =='psl':
            sheet['L4'] = int(sheet['L4'].value) + 1
        elif cell.value == 'sl':
            sheet['N4'] = int(sheet['N4'].value) + 1
        zliczanie += 1

sheet['G4'] = zliczanie
workbench.save("Imienne_Rozliczenie_Bojowe2.xlsx")