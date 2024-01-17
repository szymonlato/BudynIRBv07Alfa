import openpyxl
from pathlib import Path

pathToIRB = "EmptyFiles/Imienne_Rozliczenie_Bojowe.xlsx"


def createSheet(path):
    xlsxFile = Path(path)
    wbObj = openpyxl.load_workbook(xlsxFile)
    return wbObj

powiadomienieOKoncuPrzepustki = ["00.33.00", "05.15.00"]

def SprawdzenieCzyKoniecPrzepustki():
    workbench = createSheet(pathToIRB)
    sheet = workbench.active
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=6)):
        if sheet.cell(i + 3, 5).value == 'PS':
            print("wow dziala")
            #threading.Thread(target=KomunikatKtoNieWrocil, args=(sheet.cell(i + 3, 2).value,)).start()

def wypelnijComboBox_3():
    #self.ui.comboBox_3.clear()
    #self.ui.comboBox_3.addItem("--", 0)
    workbench = createSheet(pathToIRB)
    sheet = workbench.active
    zliczanie = 1
    for i, row in enumerate(sheet.iter_rows(min_row=3, max_row=sheet.max_row, min_col=5, max_col=6)):
        if sheet.cell(i + 3, 5).value == None and sheet.cell(i + 3, 5).value != "l4":
            #print(sheet.cell(i + 3, 3).value)
            self.ui.comboBox_3.addItem(sheet.cell(i + 3, 3).value, zliczanie)
            zliczanie += 1

#SprawdzenieCzyKoniecPrzepustki()
wypelnijComboBox_3()